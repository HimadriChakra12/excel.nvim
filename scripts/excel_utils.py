#!/usr/bin/env python3
"""
Excel.nvim utility script for advanced Excel operations
"""

import sys
import json
import argparse
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(json.dumps({
        'error': f'Missing required package: {e.name}',
        'message': 'Install with: pip install pandas openpyxl'
    }))
    sys.exit(1)


def list_sheets(excel_file):
    """List all sheets in an Excel file"""
    try:
        xlsx = pd.ExcelFile(excel_file)
        return {
            'sheets': xlsx.sheet_names,
            'count': len(xlsx.sheet_names)
        }
    except Exception as e:
        return {'error': str(e)}


def sheet_info(excel_file, sheet_name=None):
    """Get information about a specific sheet"""
    try:
        wb = load_workbook(excel_file, data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        return {
            'name': ws.title,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'dimensions': ws.dimensions,
        }
    except Exception as e:
        return {'error': str(e)}


def convert_to_csv(excel_file, output_csv, sheet_index=0):
    """Convert Excel sheet to CSV"""
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_index)
        df.to_csv(output_csv, index=False)
        return {
            'success': True,
            'rows': len(df),
            'columns': len(df.columns),
            'output': output_csv
        }
    except Exception as e:
        return {'error': str(e)}


def csv_to_excel(csv_file, excel_file, sheet_name='Sheet1', sheet_index=0):
    """Update Excel file from CSV"""
    try:
        df = pd.read_csv(csv_file)
        
        # Load existing workbook if it exists
        if Path(excel_file).exists():
            wb = load_workbook(excel_file)
            if sheet_index < len(wb.sheetnames):
                sheet_name = wb.sheetnames[sheet_index]
                ws = wb[sheet_name]
                # Clear existing data
                ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet(sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
        
        # Write data
        for row_idx, row in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(excel_file)
        return {
            'success': True,
            'rows': len(df) + 1,  # +1 for header
            'columns': len(df.columns)
        }
    except Exception as e:
        return {'error': str(e)}


def create_workbook(excel_file, sheet_name='Sheet1'):
    """Create a new Excel workbook"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add default headers
        headers = ['Column1', 'Column2', 'Column3']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        
        wb.save(excel_file)
        return {
            'success': True,
            'file': excel_file,
            'sheet': sheet_name
        }
    except Exception as e:
        return {'error': str(e)}


def add_formula(excel_file, sheet_index, cell_ref, formula):
    """Add a formula to a specific cell"""
    try:
        wb = load_workbook(excel_file)
        sheet_name = wb.sheetnames[sheet_index]
        ws = wb[sheet_name]
        
        ws[cell_ref] = formula
        
        wb.save(excel_file)
        return {
            'success': True,
            'cell': cell_ref,
            'formula': formula
        }
    except Exception as e:
        return {'error': str(e)}


def format_cell(excel_file, sheet_index, cell_ref, bold=False, color=None, bg_color=None):
    """Format a specific cell"""
    try:
        wb = load_workbook(excel_file)
        sheet_name = wb.sheetnames[sheet_index]
        ws = wb[sheet_name]
        
        cell = ws[cell_ref]
        
        if bold:
            cell.font = Font(bold=True)
        if color:
            cell.font = Font(color=color)
        if bg_color:
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        
        wb.save(excel_file)
        return {
            'success': True,
            'cell': cell_ref
        }
    except Exception as e:
        return {'error': str(e)}


def main():
    parser = argparse.ArgumentParser(description='Excel.nvim utility script')
    parser.add_argument('command', choices=[
        'list-sheets', 'sheet-info', 'to-csv', 'from-csv',
        'create', 'add-formula', 'format-cell'
    ])
    parser.add_argument('--file', required=True, help='Excel file path')
    parser.add_argument('--csv', help='CSV file path')
    parser.add_argument('--sheet', help='Sheet name or index', default=0, type=int)
    parser.add_argument('--cell', help='Cell reference (e.g., A1)')
    parser.add_argument('--formula', help='Formula to add')
    parser.add_argument('--bold', action='store_true', help='Bold text')
    parser.add_argument('--color', help='Font color (hex)')
    parser.add_argument('--bg-color', help='Background color (hex)')
    
    args = parser.parse_args()
    
    result = None
    
    if args.command == 'list-sheets':
        result = list_sheets(args.file)
    elif args.command == 'sheet-info':
        result = sheet_info(args.file, args.sheet)
    elif args.command == 'to-csv':
        if not args.csv:
            result = {'error': '--csv argument required'}
        else:
            result = convert_to_csv(args.file, args.csv, args.sheet)
    elif args.command == 'from-csv':
        if not args.csv:
            result = {'error': '--csv argument required'}
        else:
            result = csv_to_excel(args.csv, args.file, sheet_index=args.sheet)
    elif args.command == 'create':
        result = create_workbook(args.file)
    elif args.command == 'add-formula':
        if not args.cell or not args.formula:
            result = {'error': '--cell and --formula arguments required'}
        else:
            result = add_formula(args.file, args.sheet, args.cell, args.formula)
    elif args.command == 'format-cell':
        if not args.cell:
            result = {'error': '--cell argument required'}
        else:
            result = format_cell(args.file, args.sheet, args.cell, 
                               args.bold, args.color, args.bg_color)
    
    print(json.dumps(result, indent=2))


if __name__ == '__main__':
    main()
