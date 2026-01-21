#!/usr/bin/env python3
"""
Create sample Excel files for testing excel.nvim
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    exit(1)

import os

def create_sample_budget():
    """Create a sample budget spreadsheet"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget 2025"
    
    # Headers
    ws['A1'] = 'Category'
    ws['B1'] = 'January'
    ws['C1'] = 'February'
    ws['D1'] = 'March'
    ws['E1'] = 'Total'
    
    # Make headers bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Data
    categories = [
        ('Rent', 1200, 1200, 1200),
        ('Groceries', 450, 480, 460),
        ('Utilities', 150, 160, 155),
        ('Transportation', 200, 220, 210),
        ('Entertainment', 100, 120, 110),
    ]
    
    row = 2
    for category, jan, feb, mar in categories:
        ws[f'A{row}'] = category
        ws[f'B{row}'] = jan
        ws[f'C{row}'] = feb
        ws[f'D{row}'] = mar
        ws[f'E{row}'] = f'=SUM(B{row}:D{row})'
        row += 1
    
    # Total row
    ws[f'A{row}'] = 'Total'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = '=SUM(B2:B6)'
    ws[f'C{row}'] = '=SUM(C2:C6)'
    ws[f'D{row}'] = '=SUM(D2:D6)'
    ws[f'E{row}'] = '=SUM(E2:E6)'
    
    # Add another sheet
    ws2 = wb.create_sheet("Summary")
    ws2['A1'] = 'Total Expenses'
    ws2['B1'] = "=Budget 2025!E7"
    ws2['A1'].font = Font(bold=True)
    
    wb.save('sample_budget.xlsx')
    print("✓ Created: sample_budget.xlsx")

def create_sample_grades():
    """Create a sample gradebook"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"
    
    # Headers
    headers = ['Student', 'Homework', 'Midterm', 'Final', 'Average', 'Grade']
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(1, col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # Sample data
    students = [
        ('Alice Johnson', 95, 88, 92),
        ('Bob Smith', 87, 90, 89),
        ('Carol Davis', 92, 85, 88),
        ('David Wilson', 78, 82, 80),
        ('Eve Martinez', 90, 93, 94),
    ]
    
    for row, (name, hw, mid, final) in enumerate(students, 2):
        ws[f'A{row}'] = name
        ws[f'B{row}'] = hw
        ws[f'C{row}'] = mid
        ws[f'D{row}'] = final
        ws[f'E{row}'] = f'=(B{row}*0.3+C{row}*0.3+D{row}*0.4)'
        ws[f'F{row}'] = f'=IF(E{row}>=90,"A",IF(E{row}>=80,"B",IF(E{row}>=70,"C","F")))'
    
    # Statistics
    stats_row = len(students) + 3
    ws[f'A{stats_row}'] = 'Class Average'
    ws[f'A{stats_row}'].font = Font(bold=True)
    ws[f'E{stats_row}'] = '=AVERAGE(E2:E6)'
    
    ws.column_dimensions['A'].width = 20
    
    wb.save('sample_grades.xlsx')
    print("✓ Created: sample_grades.xlsx")

def create_sample_inventory():
    """Create a sample inventory spreadsheet"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Headers
    headers = ['Item', 'Quantity', 'Price', 'Total Value', 'Reorder?']
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
        ws.cell(1, col).font = Font(bold=True)
    
    # Data
    items = [
        ('Widget A', 50, 12.50),
        ('Widget B', 25, 8.75),
        ('Widget C', 100, 5.00),
        ('Widget D', 15, 22.00),
        ('Widget E', 75, 15.50),
    ]
    
    for row, (item, qty, price) in enumerate(items, 2):
        ws[f'A{row}'] = item
        ws[f'B{row}'] = qty
        ws[f'C{row}'] = price
        ws[f'D{row}'] = f'=B{row}*C{row}'
        ws[f'E{row}'] = f'=IF(B{row}<30,"Yes","No")'
    
    # Total
    total_row = len(items) + 2
    ws[f'A{total_row}'] = 'Total Inventory Value'
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'D{total_row}'] = f'=SUM(D2:D{len(items)+1})'
    
    wb.save('sample_inventory.xlsx')
    print("✓ Created: sample_inventory.xlsx")

def main():
    print("Creating sample Excel files for testing excel.nvim...")
    print()
    
    create_sample_budget()
    create_sample_grades()
    create_sample_inventory()
    
    print()
    print("Sample files created successfully!")
    print()
    print("Test them with:")
    print("  nvim sample_budget.xlsx")
    print("  nvim sample_grades.xlsx")
    print("  nvim sample_inventory.xlsx")

if __name__ == '__main__':
    main()
