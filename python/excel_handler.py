#!/usr/bin/env python3
"""
Excel handler for excel.nvim
Handles reading, writing, and manipulating Excel files
Supports: .xlsx, .xls, .xlsm, .xlsb, .csv
"""

import sys
import json
import os
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}))
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    pd = None


class ExcelHandler:
    def __init__(self):
        self.workbook = None
        self.filepath = None
        
    def load(self, filepath):
        """Load Excel file and return sheet data"""
        try:
            self.filepath = filepath
            
            # Handle CSV files
            if filepath.endswith('.csv'):
                return self._load_csv(filepath)
            
            # Handle Excel files
            if filepath.endswith(('.xls', '.xlsx', '.xlsm', '.xlsb')):
                return self._load_excel(filepath)
            
            return {"error": f"Unsupported file format: {filepath}"}
            
        except Exception as e:
            return {"error": f"Failed to load file: {str(e)}"}
    
    def _load_excel(self, filepath):
        """Load Excel file"""
        try:
            self.workbook = load_workbook(filepath, data_only=False)
            sheets = self.workbook.sheetnames
            current_sheet = self.workbook.active.title
            
            # Read data from all sheets
            data = {}
            for sheet_name in sheets:
                sheet = self.workbook[sheet_name]
                data[sheet_name] = self._read_sheet(sheet)
            
            return {
                "success": True,
                "sheets": sheets,
                "current_sheet": current_sheet,
                "data": data
            }
        except Exception as e:
            return {"error": f"Failed to load Excel file: {str(e)}"}
    
    def _load_csv(self, filepath):
        """Load CSV file"""
        try:
            if pd is None:
                # Fallback to basic CSV reading
                import csv
                data = {}
                with open(filepath, 'r') as f:
                    reader = csv.reader(f)
                    for row_idx, row in enumerate(reader, 1):
                        for col_idx, value in enumerate(row, 1):
                            key = f"{row_idx},{col_idx}"
                            data[key] = value
                
                return {
                    "success": True,
                    "sheets": ["Sheet1"],
                    "current_sheet": "Sheet1",
                    "data": {"Sheet1": data}
                }
            
            # Use pandas for better CSV handling
            df = pd.read_csv(filepath)
            data = {}
            
            # Add header row
            for col_idx, col_name in enumerate(df.columns, 1):
                data[f"1,{col_idx}"] = str(col_name)
            
            # Add data rows
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    if pd.notna(value):
                        data[f"{row_idx + 2},{col_idx}"] = str(value)
            
            return {
                "success": True,
                "sheets": ["Sheet1"],
                "current_sheet": "Sheet1",
                "data": {"Sheet1": data}
            }
        except Exception as e:
            return {"error": f"Failed to load CSV file: {str(e)}"}
    
    def _read_sheet(self, sheet):
        """Read all cells from a sheet"""
        data = {}
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    key = f"{cell.row},{cell.column}"
                    
                    # Handle formulas
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        data[key] = cell.value
                    else:
                        data[key] = str(cell.value)
        
        return data
    
    def save(self, filepath, data_json, current_sheet):
        """Save Excel file with updated data"""
        try:
            data = json.loads(data_json)
            
            # Handle CSV files
            if filepath.endswith('.csv'):
                return self._save_csv(filepath, data, current_sheet)
            
            # Handle Excel files
            if filepath.endswith(('.xlsx', '.xlsm', '.xlsb')):
                return self._save_excel(filepath, data, current_sheet)
            
            # Handle legacy .xls by converting to .xlsx
            if filepath.endswith('.xls'):
                new_filepath = filepath + 'x'
                result = self._save_excel(new_filepath, data, current_sheet)
                if result.get('success'):
                    result['message'] = f"Saved as {new_filepath} (converted from .xls)"
                return result
            
            return {"error": f"Unsupported file format: {filepath}"}
            
        except Exception as e:
            return {"error": f"Failed to save file: {str(e)}"}
    
    def _save_excel(self, filepath, data, current_sheet):
        """Save Excel file"""
        try:
            # Create new workbook or use existing
            if self.workbook:
                wb = self.workbook
            else:
                wb = Workbook()
                # Remove default sheet
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # Update or create sheets
            for sheet_name, sheet_data in data.items():
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    # Clear existing data
                    sheet.delete_rows(1, sheet.max_row)
                else:
                    sheet = wb.create_sheet(sheet_name)
                
                # Write data
                for cell_key, value in sheet_data.items():
                    row, col = cell_key.split(',')
                    row, col = int(row), int(col)
                    
                    cell = sheet.cell(row=row, column=col)
                    
                    # Handle formulas
                    if isinstance(value, str) and value.startswith('='):
                        cell.value = value
                    else:
                        # Try to convert to number
                        try:
                            cell.value = float(value) if '.' in value else int(value)
                        except (ValueError, TypeError):
                            cell.value = value
            
            # Save file
            wb.save(filepath)
            
            return {
                "success": True,
                "message": f"Saved to {filepath}"
            }
        except Exception as e:
            return {"error": f"Failed to save Excel file: {str(e)}"}
    
    def _save_csv(self, filepath, data, current_sheet):
        """Save CSV file"""
        try:
            sheet_data = data.get(current_sheet, {})
            
            if not sheet_data:
                return {"error": "No data to save"}
            
            # Find dimensions
            max_row = 0
            max_col = 0
            
            for cell_key in sheet_data.keys():
                row, col = cell_key.split(',')
                row, col = int(row), int(col)
                max_row = max(max_row, row)
                max_col = max(max_col, col)
            
            # Create CSV data
            import csv
            with open(filepath, 'w', newline='') as f:
                writer = csv.writer(f)
                
                for row in range(1, max_row + 1):
                    row_data = []
                    for col in range(1, max_col + 1):
                        cell_key = f"{row},{col}"
                        value = sheet_data.get(cell_key, '')
                        row_data.append(value)
                    writer.writerow(row_data)
            
            return {
                "success": True,
                "message": f"Saved to {filepath}"
            }
        except Exception as e:
            return {"error": f"Failed to save CSV file: {str(e)}"}
    
    def recalc(self, filepath):
        """Recalculate formulas using LibreOffice"""
        try:
            # Check if LibreOffice is available
            import subprocess
            
            # Try to find LibreOffice
            libreoffice_cmd = None
            for cmd in ['libreoffice', 'soffice']:
                try:
                    subprocess.run([cmd, '--version'], 
                                 capture_output=True, 
                                 timeout=5)
                    libreoffice_cmd = cmd
                    break
                except (FileNotFoundError, subprocess.TimeoutExpired):
                    continue
            
            if not libreoffice_cmd:
                return {
                    "success": False,
                    "error": "LibreOffice not found. Install it to recalculate formulas."
                }
            
            # Convert file to have LibreOffice recalculate
            abs_path = os.path.abspath(filepath)
            
            cmd = [
                libreoffice_cmd,
                '--headless',
                '--convert-to', 'xlsx',
                '--outdir', os.path.dirname(abs_path),
                abs_path
            ]
            
            result = subprocess.run(cmd, 
                                  capture_output=True, 
                                  text=True, 
                                  timeout=30)
            
            if result.returncode == 0:
                return {
                    "success": True,
                    "message": "Formulas recalculated successfully"
                }
            else:
                return {
                    "success": False,
                    "error": f"LibreOffice error: {result.stderr}"
                }
                
        except Exception as e:
            return {
                "success": False,
                "error": f"Failed to recalculate formulas: {str(e)}"
            }


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No action specified"}))
        sys.exit(1)
    
    action = sys.argv[1]
    handler = ExcelHandler()
    
    if action == 'load':
        if len(sys.argv) < 3:
            print(json.dumps({"error": "No filepath specified"}))
            sys.exit(1)
        
        filepath = sys.argv[2]
        result = handler.load(filepath)
        print(json.dumps(result))
    
    elif action == 'save':
        if len(sys.argv) < 5:
            print(json.dumps({"error": "Invalid save arguments"}))
            sys.exit(1)
        
        filepath = sys.argv[2]
        data_json = sys.argv[3]
        current_sheet = sys.argv[4]
        
        result = handler.save(filepath, data_json, current_sheet)
        print(json.dumps(result))
    
    elif action == 'recalc':
        if len(sys.argv) < 3:
            print(json.dumps({"error": "No filepath specified"}))
            sys.exit(1)
        
        filepath = sys.argv[2]
        result = handler.recalc(filepath)
        print(json.dumps(result))
    
    else:
        print(json.dumps({"error": f"Unknown action: {action}"}))
        sys.exit(1)


if __name__ == '__main__':
    main()
