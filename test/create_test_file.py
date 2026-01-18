#!/usr/bin/env python3
"""
Test script for excel.nvim
Creates a sample Excel file for testing
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Create sample data
data = {
    'Product': ['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Headphones'],
    'Price': [999.99, 29.99, 79.99, 299.99, 149.99],
    'Quantity': [5, 15, 10, 8, 12],
    'Total': ['=B2*C2', '=B3*C3', '=B4*C4', '=B5*C5', '=B6*C6']
}

df = pd.DataFrame(data)

# Create workbook using openpyxl for formulas
wb = Workbook()
ws = wb.active
ws.title = 'Sales'

# Add headers with formatting
headers = ['Product', 'Price', 'Quantity', 'Total']
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# Add data
for row_idx in range(len(df)):
    ws.cell(row=row_idx+2, column=1, value=df.iloc[row_idx]['Product'])
    ws.cell(row=row_idx+2, column=2, value=df.iloc[row_idx]['Price'])
    ws.cell(row=row_idx+2, column=3, value=df.iloc[row_idx]['Quantity'])
    ws.cell(row=row_idx+2, column=4, value=df.iloc[row_idx]['Total'])

# Add totals row
total_row = len(df) + 2
ws.cell(row=total_row, column=1, value='TOTAL')
ws.cell(row=total_row, column=1).font = Font(bold=True)
ws.cell(row=total_row, column=4, value=f'=SUM(D2:D{len(df)+1})')
ws.cell(row=total_row, column=4).font = Font(bold=True)

# Create second sheet
ws2 = wb.create_sheet('Inventory')
ws2['A1'] = 'Item'
ws2['B1'] = 'Stock'
ws2['C1'] = 'Reorder Level'

inventory_data = [
    ['Laptop', 5, 10],
    ['Mouse', 15, 20],
    ['Keyboard', 10, 15],
    ['Monitor', 8, 12],
    ['Headphones', 12, 15],
]

for row_idx, row in enumerate(inventory_data, start=2):
    for col_idx, value in enumerate(row, start=1):
        ws2.cell(row=row_idx, column=col_idx, value=value)

# Format headers
for cell in ws2[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')

# Save the workbook
output_file = 'test_workbook.xlsx'
wb.save(output_file)

print(f"Created test Excel file: {output_file}")
print(f"Sheets: {wb.sheetnames}")
print(f"\nTo test in Neovim:")
print(f"  :ExcelOpen {output_file}")
print(f"  :ExcelSheets")
print(f"  :ExcelSwitchSheet 1")
print(f"  :ExcelSave")
